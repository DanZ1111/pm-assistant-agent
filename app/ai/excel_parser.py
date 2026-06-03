"""Build 30B — workbook (xlsx / xlsm / xls / csv) → sheet-aware plain text.

The AI extractor at `app.ai.parser.extract_batch_from_workbook_text` is
intentionally text-in / json-out so it works equally well across all four
file types. This module just normalises the file into a string the prompt
can reason about, with sheet + row markers so the AI can populate the
`source_sheet` and `source_row_hint` fields per extracted project.

Hard cap on output size to keep token usage bounded.
"""
from __future__ import annotations

import csv as _csv
import io
import os
from pathlib import Path

WORKBOOK_TEXT_CAP_CHARS = 100_000
# Per-cell guard so a single absurdly-long cell can't crowd everything else out.
_PER_CELL_TRUNC = 500


def detect_workbook_kind(file_path: str) -> str:
    """Map a file path to one of: 'xlsx' | 'xls' | 'csv' | 'unsupported'.

    Mirrors the extension classification in app.routes.files.FILE_TYPE_MAP
    but narrower (only the workbook formats this parser knows how to read).
    """
    ext = file_path.rsplit(".", 1)[-1].lower() if "." in file_path else ""
    if ext in ("xlsx", "xlsm"):
        return "xlsx"
    if ext == "xls":
        return "xls"
    if ext == "csv":
        return "csv"
    return "unsupported"


def _truncate_cell(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ").strip()
    if len(text) > _PER_CELL_TRUNC:
        return text[:_PER_CELL_TRUNC] + "…(truncated)"
    return text


def _row_text(row_values, row_num: int) -> str:
    cells = [_truncate_cell(v) for v in row_values]
    if not any(cells):
        return ""  # empty row — skip
    # Join with " | " so the AI can spot column boundaries without us
    # imposing a fixed schema. Header rows look the same as data rows.
    body = " | ".join(cells)
    return f"  row {row_num}: {body}"


def _read_xlsx(file_path: str) -> list[tuple[str, list[str]]]:
    """Return [(sheet_name, [row_text, ...]), ...] for .xlsx / .xlsm files."""
    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheets: list[tuple[str, list[str]]] = []
    for ws in wb.worksheets:
        rows: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            line = _row_text(row, i)
            if line:
                rows.append(line)
        sheets.append((ws.title, rows))
    wb.close()
    return sheets


def _read_xls(file_path: str) -> list[tuple[str, list[str]]]:
    """Return sheets for .xls (legacy Excel 97-2003) via xlrd<2.0."""
    import xlrd
    book = xlrd.open_workbook(file_path)
    sheets: list[tuple[str, list[str]]] = []
    for ws in book.sheets():
        rows: list[str] = []
        for i in range(ws.nrows):
            line = _row_text(ws.row_values(i), i + 1)
            if line:
                rows.append(line)
        sheets.append((ws.name, rows))
    return sheets


def _read_csv(file_path: str) -> list[tuple[str, list[str]]]:
    """Return a single 'CSV' pseudo-sheet for .csv files."""
    rows: list[str] = []
    # Try utf-8 then fall back to latin-1 so we don't bail on legacy encodings.
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(file_path, "r", encoding=encoding, newline="") as fh:
                # Sniff dialect; fall back to comma if sniff fails.
                sample = fh.read(8192)
                fh.seek(0)
                try:
                    dialect = _csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except _csv.Error:
                    dialect = _csv.excel
                reader = _csv.reader(fh, dialect)
                for i, row in enumerate(reader, 1):
                    line = _row_text(row, i)
                    if line:
                        rows.append(line)
            break
        except UnicodeDecodeError:
            rows.clear()
            continue
    return [("CSV", rows)]


def workbook_to_text(file_path: str) -> str:
    """Render a workbook as labelled plain text. Caps at WORKBOOK_TEXT_CAP_CHARS.

    Raises ValueError for unsupported extensions; caller should detect first
    via detect_workbook_kind() and surface a friendly error to the user.
    """
    kind = detect_workbook_kind(file_path)
    if kind == "xlsx":
        sheets = _read_xlsx(file_path)
    elif kind == "xls":
        sheets = _read_xls(file_path)
    elif kind == "csv":
        sheets = _read_csv(file_path)
    else:
        raise ValueError(f"unsupported workbook extension: {file_path}")

    buf = io.StringIO()
    for name, rows in sheets:
        if not rows:
            continue
        buf.write(f"Sheet: {name}\n")
        for line in rows:
            buf.write(line)
            buf.write("\n")
        buf.write("\n")
    text = buf.getvalue()
    if len(text) > WORKBOOK_TEXT_CAP_CHARS:
        # Hard cap — keep the AI prompt manageable. Caller surfaces this.
        return ""
    return text


def extract_from_workbook(file_path: str) -> dict:
    """Top-level entry point for excel/csv intake.

    Mirrors the shape of extract_from_pdf / extract_from_image:
      success → {"workbook_text": "...", "kind": "xlsx" | "xls" | "csv"}
      failure → {"_error": "..."}
    The AI step (text → projects JSON) happens separately so test_build30b
    can exercise the parser alone without burning tokens.
    """
    kind = detect_workbook_kind(file_path)
    if kind == "unsupported":
        return {"_error": f"Unsupported workbook extension: {Path(file_path).name}"}
    try:
        text = workbook_to_text(file_path)
    except Exception as exc:
        return {"_error": f"Could not parse workbook: {exc}"}
    if not text:
        return {"_error": f"Workbook exceeds {WORKBOOK_TEXT_CAP_CHARS:,}-char cap — please split into smaller files."}
    return {"workbook_text": text, "kind": kind}
